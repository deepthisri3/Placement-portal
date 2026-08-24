"""
Report service — branch-wise placement statistics.

Fixed logic:
  total_offers         = COUNT of placement_record rows WHERE student_id IS NOT NULL
                         (one row per student per company = one offer)
  unique_placed_students = COUNT(DISTINCT student_id) WHERE student_id IS NOT NULL
                         (one student with 3 offers = 3 total_offers, 1 unique student)

Before this fix, total_offers counted ALL rows including unlinked historical
records (student_id = NULL), inflating the number massively.
"""
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.academic import Branch, Batch
from app.models.placement_record import PlacementRecord
from app.models.student import Student


def _normalize_branch(value: Optional[str]) -> str:
    return (value or "").strip().upper()


def get_branch_placement_statistics(
    db: Session,
    branch: Optional[str] = None,
    graduation_year: Optional[int] = None,
) -> List[Dict[str, Any]]:

    # ── Which branches to report on ──────────────────────────────────────────
    branch_query = db.query(Branch).order_by(Branch.code.asc())
    if branch:
        branch_query = branch_query.filter(
            func.lower(Branch.code) == branch.strip().lower()
        )
    managed_branches = branch_query.all()
    if not managed_branches:
        return []

    # ── Total students per branch ─────────────────────────────────────────────
    total_q = (
        db.query(
            func.upper(Student.branch).label("branch"),
            func.count(Student.id).label("cnt"),
        )
        .filter(Student.branch.isnot(None))
    )

    # ── Eligible students per branch ─────────────────────────────────────────
    eligible_q = (
        db.query(
            func.upper(Student.branch).label("branch"),
            func.count(Student.id).label("cnt"),
        )
        .filter(
            Student.branch.isnot(None),
            Student.ssc_percentage >= 60,
            Student.intermediate_percentage >= 60,
            Student.cgpa >= 6.5,
        )
    )

    # ── Total offers per branch ───────────────────────────────────────────────
    # Only count rows that are LINKED to a real student (student_id IS NOT NULL).
    # One row = one offer. A student with 3 offers has 3 rows → total_offers = 3.
    offers_q = (
        db.query(
            func.upper(PlacementRecord.branch).label("branch"),
            func.count(PlacementRecord.id).label("cnt"),
        )
        .filter(
            PlacementRecord.branch.isnot(None),
            PlacementRecord.student_id.isnot(None),   # ← KEY FIX
        )
    )

    # ── Unique placed students per branch ─────────────────────────────────────
    # COUNT(DISTINCT student_id) so a student with 3 offers counts as 1.
    unique_q = (
        db.query(
            func.upper(PlacementRecord.branch).label("branch"),
            func.count(func.distinct(PlacementRecord.student_id)).label("cnt"),
        )
        .filter(
            PlacementRecord.branch.isnot(None),
            PlacementRecord.student_id.isnot(None),
        )
    )

    # ── Apply year filter if provided ─────────────────────────────────────────
    if graduation_year is not None:
        total_q    = total_q.filter(Student.graduation_year == graduation_year)
        eligible_q = eligible_q.filter(Student.graduation_year == graduation_year)
        offers_q   = offers_q.filter(PlacementRecord.graduation_year == graduation_year)
        unique_q   = unique_q.filter(PlacementRecord.graduation_year == graduation_year)

    # ── Execute ───────────────────────────────────────────────────────────────
    totals_by_branch  = {
        _normalize_branch(b): int(c or 0)
        for b, c in total_q.group_by(func.upper(Student.branch)).all()
    }
    eligible_by_branch = {
        _normalize_branch(b): int(c or 0)
        for b, c in eligible_q.group_by(func.upper(Student.branch)).all()
    }
    offers_by_branch = {
        _normalize_branch(b): int(c or 0)
        for b, c in offers_q.group_by(func.upper(PlacementRecord.branch)).all()
    }
    unique_by_branch = {
        _normalize_branch(b): int(c or 0)
        for b, c in unique_q.group_by(func.upper(PlacementRecord.branch)).all()
    }

    # ── Build response rows ───────────────────────────────────────────────────
    rows: List[Dict[str, Any]] = []
    for mb in managed_branches:
        code = _normalize_branch(mb.code)
        total_s  = totals_by_branch.get(code, 0)
        unique_p = unique_by_branch.get(code, 0)
        rows.append({
            "branch":                 code,
            "total_students":         total_s,
            "eligible_students":      eligible_by_branch.get(code, 0),
            "total_offers":           offers_by_branch.get(code, 0),
            "unique_placed_students": unique_p,
            # Convenience percentage for the frontend
            "placement_percentage": (
                round(unique_p / total_s * 100, 1) if total_s > 0 else 0.0
            ),
        })

    return rows


def export_branch_placement_statistics_excel(
    db: Session,
    branch: Optional[str] = None,
    graduation_year: Optional[int] = None,
) -> StreamingResponse:
    report_rows = get_branch_placement_statistics(
        db, branch=branch, graduation_year=graduation_year
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Branch Placement Statistics"

    ws.merge_cells("A1:F1")
    ws["A1"] = "Branch-wise Placement Statistics Report"
    ws["A1"].font = Font(size=18, bold=True, color="1F2937")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = "Eligibility rule: SSC ≥ 60%, Intermediate ≥ 60%, B.Tech CGPA ≥ 6.5 | Only verified (linked) students counted"
    ws["A2"].font = Font(size=11, italic=True, color="475569")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.append([])

    headers = [
        "Branch",
        "Total Students",
        "Eligible (SSC ≥ 60%, Inter ≥ 60%, CGPA ≥ 6.5)",
        "Total Offers",
        "Unique Placed Students",
        "Placement %",
    ]
    ws.append(headers)

    for row in report_rows:
        ws.append([
            row["branch"],
            row["total_students"],
            row["eligible_students"],
            row["total_offers"],
            row["unique_placed_students"],
            f"{row['placement_percentage']}%",
        ])

    header_fill = PatternFill(fill_type="solid", fgColor="E8EEF9")
    thin = Border(
        left=Side(style="thin", color="B8C2D2"),
        right=Side(style="thin", color="B8C2D2"),
        top=Side(style="thin", color="B8C2D2"),
        bottom=Side(style="thin", color="B8C2D2"),
    )

    for cell in ws[4]:
        cell.font = Font(bold=True, color="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
        cell.fill = header_fill

    for idx, row in enumerate(
        ws.iter_rows(min_row=5, max_row=ws.max_row), start=0
    ):
        fill_color = "F8FAFC" if idx % 2 == 0 else "FFFFFF"
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)

    for column in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in column if c.value is not None),
            default=10,
        )
        ws.column_dimensions[column[0].column_letter].width = min(max_len + 3, 48)

    ws.freeze_panes = "A5"
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"branch_placement_statistics_{timestamp}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )