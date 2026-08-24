"""
Business logic for placement records.

Key rules implemented here:
  1. Only rows where Status = "Selected" (or equivalent) are saved.
     Rows with any other status (Rejected, Not Placed, etc.) are skipped.
  2. student_id is resolved by matching roll_number -> students.register_number
     (case-insensitive). If no match, the row is SKIPPED entirely — we no
     longer save unlinked records because they corrupt the statistics.
  3. Dedup key is (roll_number, company_id). Same student at same company
     upserts the row. Same student at different companies = multiple rows
     = multiple offers, but the report counts them as one unique student.
  4. Company names are resolved/created automatically.
"""
import csv
import io
from datetime import date
from typing import Dict, List, Optional, Tuple

from fastapi import HTTPException
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.company import Company
from app.models.placement_record import PlacementRecord

# ── Header aliases ────────────────────────────────────────────────────────────
# Maps every plausible column heading in the uploaded file to our internal name.
# Your Excel uses: Register Number, Student Name, Department, Company Name,
#                  Role, Package, Status

HEADER_ALIASES: Dict[str, set] = {
    "roll_number": {
        "roll number", "rollnumber", "roll no", "rollno",
        "register number", "registernumber", "reg no", "reg number",
        "registration number", "student id", "studentid", "regno",
    },
    "student_name": {
        "student name", "name", "studentname", "full name", "fullname",
    },
    "branch": {
        "branch", "department", "dept", "branch code",
    },
    "graduation_year": {
        "graduation year", "grad year", "passing year", "year",
        "graduationyear", "batch", "passout year",
    },
    "company": {
        "company", "company name", "companyname", "employer",
        "organisation", "organization",
    },
    "role": {
        "role", "designation", "position", "job role", "job title",
        "jobrole", "profile",
    },
    "package": {
        "package", "ctc", "package lpa", "salary",
        "package (lpa)", "ctc (lpa)", "lpa",
    },
    "placement_date": {
        "placement date", "date", "offer date", "placementdate",
        "joining date",
    },
    "status": {
        "status", "placement status", "offer status", "result",
        "selection status",
    },
}

REQUIRED_FIELDS = {"roll_number", "company"}
ALLOWED_EXTENSIONS = {".csv", ".xlsx"}

# Values that mean the student was selected / placed
SELECTED_VALUES = {
    "selected", "placed", "offer", "offered", "yes",
    "accepted", "confirmed", "joined",
}


def _normalize_header(raw: str) -> Optional[str]:
    key = (raw or "").strip().lower()
    for field, aliases in HEADER_ALIASES.items():
        if key == field or key in aliases:
            return field
    return None


# ── File parsing ──────────────────────────────────────────────────────────────

def parse_file(filename: str, content: bytes) -> List[Dict[str, str]]:
    lower = (filename or "").lower()
    ext = lower[lower.rfind("."):] if "." in lower else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext}'. Upload a .csv or .xlsx file.",
        )
    rows = _parse_csv(content) if ext == ".csv" else _parse_xlsx(content)
    if not rows:
        raise HTTPException(status_code=400, detail="The file has no data rows.")
    return rows


def _parse_csv(content: bytes) -> List[Dict[str, str]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return _rows_to_dicts([r for r in reader])


def _parse_xlsx(content: bytes) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed. Run: pip install openpyxl",
        )
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(["" if c is None else str(c) for c in row])
    wb.close()
    return _rows_to_dicts(all_rows)


def _rows_to_dicts(all_rows: List[List[str]]) -> List[Dict[str, str]]:
    header_idx = None
    header_map = []
    for i, row in enumerate(all_rows):
        mapped = [_normalize_header(str(c)) for c in row]
        if any(m in REQUIRED_FIELDS for m in mapped):
            header_idx = i
            header_map = mapped
            break

    if header_idx is None:
        raise HTTPException(
            status_code=400,
            detail=(
                "Could not find required columns. The file must have at least "
                "a Register Number column and a Company Name column."
            ),
        )

    result: List[Dict[str, str]] = []
    for row in all_rows[header_idx + 1:]:
        if all(str(c).strip() == "" for c in row):
            continue
        record: Dict[str, str] = {}
        for col_idx, field in enumerate(header_map):
            if field and col_idx < len(row):
                record[field] = str(row[col_idx]).strip()
        result.append(record)
    return result


# ── Row validation & coercion ─────────────────────────────────────────────────

def _validate_and_coerce(raw: Dict[str, str]) -> Tuple[Optional[dict], Optional[str]]:
    """
    Returns (clean_dict, None) on success, (None, reason_string) on failure.

    Important: if a 'status' column is present and its value is NOT one of
    the SELECTED_VALUES, the row is skipped. This ensures only placed/selected
    students appear in the database.
    """
    roll = (raw.get("roll_number") or "").strip()
    company_name = (raw.get("company") or "").strip()

    if not roll:
        return None, "Missing register number — row skipped."
    if not company_name:
        return None, "Missing company name — row skipped."

    # Status check — skip if status column exists but student was not selected
    status_raw = (raw.get("status") or "").strip().lower()
    if status_raw and status_raw not in SELECTED_VALUES:
        return None, f"Status is '{raw.get('status', '')}' (not Selected) — row skipped."

    clean = {
        "roll_number":     roll,
        "company_name":    company_name,
        "student_name":    (raw.get("student_name") or "").strip() or None,
        "branch":          (raw.get("branch") or "").strip() or None,
        "role":            (raw.get("role") or "").strip() or None,
        "graduation_year": None,
        "package":         None,
        "placement_date":  None,
    }

    gy = (raw.get("graduation_year") or "").strip()
    if gy:
        try:
            clean["graduation_year"] = int(float(gy))
        except ValueError:
            return None, f"Invalid graduation year '{gy}' — row skipped."

    pkg = (raw.get("package") or "").strip().replace(",", "")
    if pkg:
        try:
            clean["package"] = float(pkg)
        except ValueError:
            return None, f"Invalid package '{pkg}' — row skipped."

    pd_raw = (raw.get("placement_date") or "").strip()
    if pd_raw:
        parsed = _parse_date(pd_raw)
        if parsed is None:
            return None, f"Invalid placement date '{pd_raw}' — row skipped."
        clean["placement_date"] = parsed

    return clean, None


def _parse_date(value: str) -> Optional[date]:
    from datetime import datetime
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


# ── Merge (upsert) ────────────────────────────────────────────────────────────

def merge_records(db: Session, rows: List[Dict[str, str]]) -> dict:
    """
    Process all rows in a single transaction:
      - Validate each row (skip invalid ones, collect reasons)
      - Skip rows where student is not "Selected"
      - Skip rows where register number doesn't match any student in our DB
      - Upsert by (roll_number, company_id)
    """
    inserted = 0
    updated = 0
    skipped = 0
    unmatched = 0
    errors = []
    companies_created = []

    company_cache: Dict[str, Company] = {}
    student_cache: Dict[str, Optional[int]] = {}

    try:
        for i, raw in enumerate(rows, start=1):
            clean, reason = _validate_and_coerce(raw)
            if reason:
                skipped += 1
                errors.append({"row": i, "reason": reason})
                continue

            # ── Resolve student FIRST — skip if not found ──────────────────
            roll = clean["roll_number"]
            if roll not in student_cache:
                student_cache[roll] = _lookup_student_id(db, roll)

            student_id = student_cache[roll]
            if student_id is None:
                unmatched += 1
                errors.append({
                    "row": i,
                    "reason": (
                        f"Register number '{roll}' not found in the student database. "
                        "This record was NOT saved. Register the student first, then "
                        "re-upload."
                    ),
                })
                continue

            # ── Resolve (or create) company by name ────────────────────────
            cname = clean["company_name"]
            company = company_cache.get(cname.lower())
            if company is None:
                company = (
                    db.query(Company)
                    .filter(func.lower(Company.name) == cname.lower())
                    .first()
                )
                if company is None:
                    company = Company(name=cname)
                    db.add(company)
                    db.flush()
                    companies_created.append(cname)
                company_cache[cname.lower()] = company

            # ── Upsert by (roll_number, company_id) ────────────────────────
            existing = (
                db.query(PlacementRecord)
                .filter(
                    PlacementRecord.roll_number == roll,
                    PlacementRecord.company_id == company.id,
                )
                .first()
            )

            if existing:
                existing.student_id     = student_id
                existing.student_name   = clean["student_name"]
                existing.branch         = clean["branch"]
                existing.graduation_year = clean["graduation_year"]
                existing.role           = clean["role"]
                existing.package        = clean["package"]
                existing.placement_date = clean["placement_date"]
                updated += 1
            else:
                db.add(PlacementRecord(
                    company_id      = company.id,
                    student_id      = student_id,
                    roll_number     = roll,
                    student_name    = clean["student_name"],
                    branch          = clean["branch"],
                    graduation_year = clean["graduation_year"],
                    role            = clean["role"],
                    package         = clean["package"],
                    placement_date  = clean["placement_date"],
                ))
                inserted += 1

        db.commit()

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Upload failed and was rolled back. No records were saved. ({e})",
        )

    return {
        "inserted":          inserted,
        "updated":           updated,
        "skipped":           skipped,
        "unmatched":         unmatched,
        "total_rows":        len(rows),
        "companies_created": companies_created,
        "errors":            errors,
    }


def _lookup_student_id(db: Session, roll_number: str) -> Optional[int]:
    """Match register number case-insensitively."""
    from app.models.student import Student
    student = (
        db.query(Student)
        .filter(func.lower(Student.register_number) == roll_number.strip().lower())
        .first()
    )
    return student.id if student else None


# ── Retrieval ─────────────────────────────────────────────────────────────────

def get_records(
    db: Session,
    year: Optional[int] = None,
    branch: Optional[str] = None,
    company_id: Optional[int] = None,
    page: int = 1,
    page_size: int = 25,
) -> dict:
    page = max(page, 1)
    page_size = min(max(page_size, 1), 200)

    q = db.query(PlacementRecord)
    if year is not None:
        q = q.filter(PlacementRecord.graduation_year == year)
    if branch:
        q = q.filter(func.lower(PlacementRecord.branch) == branch.strip().lower())
    if company_id is not None:
        q = q.filter(PlacementRecord.company_id == company_id)

    total = q.count()
    items = (
        q.order_by(
            PlacementRecord.graduation_year.desc(),
            PlacementRecord.id.desc(),
        )
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return {"total": total, "page": page, "page_size": page_size, "items": items}


def build_records_xlsx(
    db: Session,
    year: Optional[int] = None,
    branch: Optional[str] = None,
    company_id: Optional[int] = None,
) -> bytes:
    from openpyxl import Workbook
    from io import BytesIO

    q = db.query(PlacementRecord)
    if year is not None:
        q = q.filter(PlacementRecord.graduation_year == year)
    if branch:
        q = q.filter(func.lower(PlacementRecord.branch) == branch.strip().lower())
    if company_id is not None:
        q = q.filter(PlacementRecord.company_id == company_id)

    records = q.order_by(
        PlacementRecord.graduation_year.desc(),
        PlacementRecord.id.desc(),
    ).all()

    cids = {r.company_id for r in records}
    cmap: Dict[int, str] = {}
    if cids:
        for cid, cname in (
            db.query(Company.id, Company.name)
            .filter(Company.id.in_(cids))
            .all()
        ):
            cmap[cid] = cname

    wb = Workbook()
    ws = wb.active
    ws.title = "Placement Records"
    ws.append([
        "Roll Number", "Student Name", "Branch", "Graduation Year",
        "Company", "Role", "Package (LPA)", "Placement Date",
    ])
    for r in records:
        ws.append([
            r.roll_number,
            r.student_name,
            r.branch,
            r.graduation_year,
            cmap.get(r.company_id, str(r.company_id)),
            r.role,
            float(r.package) if r.package is not None else None,
            r.placement_date.isoformat() if r.placement_date else None,
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()